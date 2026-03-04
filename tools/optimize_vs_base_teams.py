#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import itertools
import json
import math
import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen


COLOR_BY_FEEL = {1: "R", 2: "B", 3: "G", 4: "Y", 5: "P"}
COLOR_WORDS = {
    "RED": "R",
    "BLUE": "B",
    "GREEN": "G",
    "YELLOW": "Y",
    "PURPLE": "P",
}
SONG_COLOR_WORDS = {
    "RED": "R",
    "BLUE": "B",
    "GREEN": "G",
    "YELLOW": "Y",
    "PURPLE": "P",
    "ALL": "ALL",
}
AXES = ("vo", "da", "pe")
# Keep fallback disabled by default.
# Some cards legitimately have no SSR+ node row in additional_card_node_training_masters
# and should stay at their base scene-card max stats (no synthetic +750 per axis).
SSR_PLUS_FALLBACK_NODE_BONUS = 0

WIKI_SAKURA_URL = "https://ja.wikipedia.org/wiki/%E6%AB%BB%E5%9D%8246"
WIKI_HINATA_URL = "https://ja.wikipedia.org/wiki/%E6%97%A5%E5%90%91%E5%9D%8246"
UNIAIR_SONGLIST_URL = "https://zawa-oden-smk.github.io/uniair_sim/songlist.js"

MANUAL_EXPECTED_OVERRIDES_RAW: dict[str, float] = {
    "100%の確率で1回だけ GOODとBADとMISS が GREAT になる": 0.0,
    "11秒おきに26%の確率で15秒間 コンボボーナス40%アップ": 3.49,
    "8秒おきに21%の確率で7秒間 GOODかBAD でもコンボが継続": 0.0,
    "9秒おきに22%の確率で8秒間 GOODかBADかMISS でもコンボが継続": 0.0,
}

NAME_NORMALIZE_TABLE = str.maketrans(
    {
        "髙": "高",
        "﨑": "崎",
        "神": "神",
        "邉": "辺",
        "濱": "浜",
        "冨": "富",
        "嶋": "島",
        "塚": "塚",
        "　": "",
        " ": "",
    }
)


def _normalize_ws(text: str) -> str:
    text = text.replace("\u00a0", " ").replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_skill_key(text: str) -> str:
    text = _normalize_ws(text)
    # Normalize minor text variants seen across masters/xlsx sheets.
    text = text.replace("コンボボーナスが", "コンボボーナス")
    text = text.replace("スコアが", "スコア")
    text = text.replace("PERFECT の", "PERFECTの")
    text = text.replace("GREAT の", "GREATの")
    text = text.replace("GOOD の", "GOODの")
    text = text.replace("BAD の", "BADの")
    text = text.replace("MISS の", "MISSの")
    return re.sub(r"\s+", "", text)


def _looks_like_skill_desc(text: str) -> bool:
    s = _normalize_ws(str(text or ""))
    if not s:
        return False
    # Reject plain title-like overrides (e.g. "Buddies") from workbook cells.
    if re.fullmatch(r"[A-Za-z0-9'’!?.:\- ]+", s):
        return False
    # Typical in-game skill text markers.
    markers = (
        "秒おきに",
        "確率",
        "スコア",
        "コンボ",
        "ライフ",
        "GREAT",
        "PERFECT",
        "GOOD",
        "BAD",
        "MISS",
        "アップ",
        "回復",
        "カット",
        "継続",
    )
    return any(m in s for m in markers)


def _normalize_name(text: str) -> str:
    return _normalize_ws(text).translate(NAME_NORMALIZE_TABLE)


def _normalize_song_name(text: str) -> str:
    return _normalize_ws(text).replace("　", "").replace(" ", "")


def _title_key(text: str) -> str:
    return _normalize_ws(text).lower().replace("é", "e")


def _is_veaut_card(card: "Card") -> bool:
    return "veaut" in _title_key(card.title)


MANUAL_EXPECTED_OVERRIDES = {
    _normalize_skill_key(k): v for k, v in MANUAL_EXPECTED_OVERRIDES_RAW.items()
}


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _find_latest_masters(roots: Path) -> Path:
    versions = sorted(p for p in roots.iterdir() if p.is_dir() and p.name.isdigit())
    if not versions:
        raise SystemExit(f"No masters versions found in {roots}")
    return versions[-1]


def _strip_tags(text: str) -> str:
    return re.sub(r"<[^>]+>", "", text or "")


def _extract_colors(text: str) -> set[str]:
    colors = set()
    for word in re.findall(r"(RED|BLUE|GREEN|YELLOW|PURPLE)", text):
        colors.add(COLOR_WORDS[word])
    return colors


def _parse_axes_text(text: str) -> set[str]:
    normalized = text.replace("、", "・").replace(",", "・")
    axes: set[str] = set()
    if "Vo" in normalized:
        axes.add("vo")
    if "Da" in normalized:
        axes.add("da")
    if "Pe" in normalized:
        axes.add("pe")
    if "全て" in normalized:
        axes.update(AXES)
    return axes


def _excel_skill_expected_map(excel_path: Path) -> dict[str, float]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise SystemExit(f"openpyxl is required: {exc}")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    out: dict[str, float] = {}

    def _expected_to_pct(value: Any) -> float | None:
        if isinstance(value, (int, float)):
            v = float(value)
            return v * 100.0 if v <= 1.0 else v
        if isinstance(value, str):
            nums = [float(x) for x in re.findall(r"([0-9]+(?:\.[0-9]+)?)", value)]
            if not nums:
                return None
            # For cells like "无特化1.35 同色2.7 满绊2.33 色绊3.68", keep the highest.
            return max(nums)
        return None

    # Primary mapping from detailed card table.
    if "欅櫻新表试用" in wb.sheetnames:
        ws = wb["欅櫻新表试用"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            ev = _expected_to_pct(row[7] if len(row) > 7 else None)  # col H: 期待值
            if ev is None:
                continue
            for idx in (8, 9):  # col I/J: 主技能 / 变化后技能
                s = row[idx] if len(row) > idx else None
                if isinstance(s, str) and s.strip():
                    out[_normalize_skill_key(s)] = ev

    # Fallback compact mapping table.
    if "卡片技能" in wb.sheetnames:
        ws = wb["卡片技能"]
        for r in range(1, ws.max_row + 1):
            short = ws.cell(r, 2).value
            full = ws.cell(r, 9).value
            if not isinstance(short, str) or not isinstance(full, str):
                continue
            m = re.search(r"[（(]\s*([0-9]+(?:\.[0-9]+)?)\s*%\s*[)）]", short)
            if not m:
                continue
            out[_normalize_skill_key(full)] = float(m.group(1))

    if not out:
        raise SystemExit(
            f"No skill->expected mapping extracted from workbook: {excel_path}"
        )
    return out


def _excel_card_after_skill_map(excel_path: Path) -> dict[tuple[str, str, str], str]:
    """
    Build per-card skill override from workbook row data.
    Key: (title_key, member_name_norm, color)
    Value: changed skill text (col J) if present, otherwise main skill text (col I).
    """
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise SystemExit(f"openpyxl is required: {exc}")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    out: dict[tuple[str, str, str], str] = {}
    if "欅櫻新表试用" not in wb.sheetnames:
        return out

    ws = wb["欅櫻新表试用"]

    def _row_member_norm(v: Any) -> str:
        s = _normalize_ws(str(v or ""))
        # Sheet member cells usually look like "23 松田里奈".
        s = re.sub(r"^\d+\s*", "", s)
        return _normalize_name(s)

    def _row_color(v: Any) -> str | None:
        s = _normalize_ws(str(v or "")).upper()
        if s in {"R", "B", "G", "Y", "P"}:
            return s
        if s in COLOR_WORDS:
            return COLOR_WORDS[s]
        return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = _normalize_ws(str(row[1] if len(row) > 1 else ""))
        member_norm = _row_member_norm(row[2] if len(row) > 2 else "")
        color = _row_color(row[3] if len(row) > 3 else "")
        if not title or not member_norm or not color:
            continue
        after_skill = row[9] if len(row) > 9 else None
        main_skill = row[8] if len(row) > 8 else None
        chosen = ""
        if isinstance(after_skill, str) and after_skill.strip():
            chosen = _normalize_ws(after_skill)
        elif isinstance(main_skill, str) and main_skill.strip():
            chosen = _normalize_ws(main_skill)
        if not chosen:
            continue
        out[(_title_key(title), member_norm, color)] = chosen

    # Additional override source for newer/full tables.
    # Note: sheets have slightly different column layouts, so we use
    # a sheet-aware + heuristic pick strategy for skill/title cells.
    def _load_full_ssr_sheet(sheet_name: str) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws_full = wb[sheet_name]
        for row in ws_full.iter_rows(min_row=2, values_only=True):
            member_norm = _row_member_norm(row[2] if len(row) > 2 else "")
            color = _row_color(row[3] if len(row) > 3 else "")
            if not member_norm or not color:
                continue

            title = ""
            # Prefer explicit normalized title column when it exists.
            title_candidates: list[Any] = []
            if "日SSR全数据" in sheet_name:
                # 日SSR全数据: V(22) title is stable.
                title_candidates.extend([row[21] if len(row) > 21 else None, row[1] if len(row) > 1 else None])
            else:
                # 欅櫻SSR全数据: S(19) is often title, B(2) also usable.
                title_candidates.extend(
                    [
                        row[18] if len(row) > 18 else None,
                        row[1] if len(row) > 1 else None,
                        row[21] if len(row) > 21 else None,
                    ]
                )

            for cand in title_candidates:
                if not (isinstance(cand, str) and cand.strip()):
                    continue
                t = _normalize_ws(cand)
                title = t.split("/", 1)[0].strip() if "/" in t else t
                if title:
                    break
            if not title:
                continue

            # Skill columns differ by sheet:
            # - 日SSR全数据: S(19) tends to be full skill text.
            # - 欅櫻SSR全数据: P(16) tends to be full skill text; S(19) is often title.
            if "日SSR全数据" in sheet_name:
                skill_candidates = [
                    row[18] if len(row) > 18 else None,  # S
                    row[19] if len(row) > 19 else None,  # T (sometimes short alias)
                    row[15] if len(row) > 15 else None,  # P fallback
                ]
            else:
                skill_candidates = [
                    row[15] if len(row) > 15 else None,  # P
                    row[16] if len(row) > 16 else None,  # Q
                    row[18] if len(row) > 18 else None,  # S (guarded by heuristic)
                ]
            chosen_skill = ""
            for cand in skill_candidates:
                if not (isinstance(cand, str) and cand.strip()):
                    continue
                cand_norm = _normalize_ws(cand)
                if _looks_like_skill_desc(cand_norm):
                    chosen_skill = cand_norm
                    break
            if not chosen_skill:
                continue

            out[(_title_key(title), member_norm, color)] = chosen_skill

    _load_full_ssr_sheet("日SSR全数据")
    _load_full_ssr_sheet("欅櫻SSR全数据")
    return out


def _excel_card_scene_upgrade_total_map(excel_path: Path) -> dict[tuple[str, str, str], int]:
    """
    Build per-card scene total upgrade map from workbook.
    Key: (title_key, member_name_norm, color)
    Value: scene total delta between Lv100 and Lv80 (usually 2250 for +750/+750/+750).
    """
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise SystemExit(f"openpyxl is required: {exc}")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    out: dict[tuple[str, str, str], int] = {}

    def _row_member_norm(v: Any) -> str:
        s = _normalize_ws(str(v or ""))
        s = re.sub(r"^\d+\s*", "", s)
        return _normalize_name(s)

    def _row_color(v: Any) -> str | None:
        s = _normalize_ws(str(v or "")).upper()
        if s in {"R", "B", "G", "Y", "P"}:
            return s
        if s in COLOR_WORDS:
            return COLOR_WORDS[s]
        return None

    def _to_int(v: Any) -> int | None:
        try:
            return int(float(v))
        except Exception:
            return None

    def _title_from_row(sheet_name: str, row: tuple[Any, ...]) -> str:
        title_candidates: list[Any] = []
        if "日SSR全数据" in sheet_name:
            title_candidates.extend([row[21] if len(row) > 21 else None, row[1] if len(row) > 1 else None])
        else:
            title_candidates.extend(
                [
                    row[18] if len(row) > 18 else None,
                    row[1] if len(row) > 1 else None,
                    row[21] if len(row) > 21 else None,
                ]
            )
        for cand in title_candidates:
            if not (isinstance(cand, str) and cand.strip()):
                continue
            t = _normalize_ws(cand)
            title = t.split("/", 1)[0].strip() if "/" in t else t
            if title:
                return title
        return ""

    for sheet_name in ("日SSR全数据", "欅櫻SSR全数据"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            member_norm = _row_member_norm(row[2] if len(row) > 2 else "")
            color = _row_color(row[3] if len(row) > 3 else "")
            title = _title_from_row(sheet_name, row)
            if not member_norm or not color or not title:
                continue

            p80 = _to_int(row[8] if len(row) > 8 else None)
            p100 = _to_int(row[9] if len(row) > 9 else None)
            if p80 is None or p100 is None:
                continue
            delta = int(p100 - p80)
            if delta <= 0:
                continue

            key = (_title_key(title), member_norm, color)
            prev = int(out.get(key, 0))
            if delta > prev:
                out[key] = delta

    return out


@dataclass(frozen=True)
class VSCenterRule:
    mode: str
    scale: float
    source_types: set[str]
    agg_target_types: set[str]
    fixed_bonus: dict[str, float]
    fixed_target_types: set[str]


@dataclass(frozen=True)
class Card:
    code: str
    title: str
    member_code: str
    member_name: str
    member_name_norm: str
    color: str
    vo: int
    da: int
    pe: int
    skill_desc: str
    skill_expected: float
    skill_is_score: bool
    skill_is_combo: bool
    leader_name: str
    leader_desc: str
    leader_effect: dict[str, float]
    leader_skill_rate_color: dict[str, float]
    leader_skill_rate_member: dict[str, float]
    is_vs_base: bool
    vs_rule: VSCenterRule | None


@dataclass(frozen=True)
class Song:
    no: int
    color: str
    name: str
    live: str
    level: int
    seconds: int
    notes: int
    psylli: float


@dataclass
class TeamResult:
    center: Card
    supports: list[Card]
    objective: float
    bench_score: int
    eff_vo: int
    eff_da: int
    eff_pe: int
    team_power: int
    score_ev_pct: float
    combo_ev_pct: float
    effective_score_ev_pct: float
    effective_combo_ev_pct: float
    score_sigma_pct: float
    combo_sigma_pct: float

    @property
    def cards(self) -> list[Card]:
        return [self.center] + self.supports


def _http_get_text(url: str) -> str:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urlopen(req, timeout=30).read().decode("utf-8", "ignore")


def _extract_current_members_from_wiki_html(html_text: str) -> list[str]:
    m_start = re.search(r'<h3 id="現メンバー"', html_text)
    if not m_start:
        return []
    section = html_text[m_start.start() :]
    m_end = re.search(r'<h3 id="(?:元メンバー|旧メンバー)"', section)
    if m_end:
        section = section[: m_end.start()]

    names: list[str] = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", section, flags=re.S):
        td = re.search(r"<td[^>]*>(.*?)</td>", tr, flags=re.S)
        if not td:
            continue
        cell = td.group(1)
        cell = re.sub(r"<sup[^>]*>.*?</sup>", "", cell, flags=re.S)
        cell = re.sub(r"<rt[^>]*>.*?</rt>", "", cell, flags=re.S)
        cell = re.sub(r"<[^>]+>", "", cell)
        name = html.unescape(cell)
        name = re.sub(r"\s+", "", name)
        name = re.sub(r"[（(].*?[）)]", "", name)
        if not name:
            continue
        if any(x in name for x in ("期生", "生年月日", "出身地", "血液型", "名前", "氏名")):
            continue
        if len(name) > 12:
            continue
        names.append(name)

    out: list[str] = []
    seen: set[str] = set()
    for n in names:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def _fetch_active_members_wiki() -> dict[str, Any]:
    sakura_html = _http_get_text(WIKI_SAKURA_URL)
    hinata_html = _http_get_text(WIKI_HINATA_URL)
    sakura = _extract_current_members_from_wiki_html(sakura_html)
    hinata = _extract_current_members_from_wiki_html(hinata_html)
    all_members = sorted(set(sakura) | set(hinata))
    return {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "sources": {
            "sakurazaka46": WIKI_SAKURA_URL,
            "hinatazaka46": WIKI_HINATA_URL,
        },
        "members": {
            "sakurazaka46": sakura,
            "hinatazaka46": hinata,
            "all": all_members,
        },
    }


def _load_active_members(path: Path, refresh: bool) -> set[str] | None:
    if refresh or not path.exists():
        data = _fetch_active_members_wiki()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if not path.exists():
        return None
    data = _read_json(path)
    members = data.get("members", {}).get("all", [])
    if not isinstance(members, list):
        return None
    names = {_normalize_name(str(x)) for x in members if str(x).strip()}
    # user explicit note
    names.add(_normalize_name("松田好花"))
    return names


def _parse_skill_rate_multiplier_text(text: str) -> float | None:
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*倍", text)
    return float(m.group(1)) if m else None


def _parse_leader_effect(name: str) -> tuple[dict[str, float], set[str]]:
    if not name:
        return ({a: 0.0 for a in AXES}, set())
    clean = name.replace("アップ", "")
    target_types: set[str] = set()
    m_target = re.search(r"<([^>]+)>", clean)
    if m_target:
        target_types = _extract_colors(m_target.group(1))
    before_pct = clean.split("%", 1)[0]
    m_pct = re.search(r"([0-9]+(?:\.[0-9]+)?)%", clean)
    if not m_pct:
        return ({a: 0.0 for a in AXES}, target_types)
    pct = float(m_pct.group(1))
    axes = _parse_axes_text(before_pct)
    if not axes:
        return ({a: 0.0 for a in AXES}, target_types)
    effect = {a: (pct if a in axes else 0.0) for a in AXES}
    return effect, target_types


def _parse_member_names_blob(blob: str) -> list[str]:
    raw = re.split(r"[・/,，、＆&]", blob)
    out: list[str] = []
    for item in raw:
        item = _normalize_name(re.sub(r"\s+", "", item))
        if item:
            out.append(item)
    return out


def _parse_leader_skill_rate_effects(name: str, description: str) -> tuple[dict[str, float], dict[str, float]]:
    text_name = _strip_tags(name)
    text_desc = _strip_tags(description)
    full = f"{text_name}\n{text_desc}"

    color_rates: dict[str, float] = {}
    member_rates: dict[str, float] = {}

    # Example: REDタイプのスキル発動率が2.25倍
    for m in re.finditer(
        r"(RED|BLUE|GREEN|YELLOW|PURPLE)\s*タイプのスキル発動率(?:が)?\s*([0-9]+(?:\.[0-9]+)?)倍",
        full,
    ):
        color = COLOR_WORDS[m.group(1)]
        val = float(m.group(2))
        color_rates[color] = max(color_rates.get(color, 1.0), val)

    # Example: スキル発動率2.0倍(森田ひかる)
    for m in re.finditer(r"スキル発動率\s*([0-9]+(?:\.[0-9]+)?)倍\(([^)]+)\)", full):
        val = float(m.group(1))
        names = _parse_member_names_blob(m.group(2))
        for nm in names:
            member_rates[nm] = max(member_rates.get(nm, 1.0), val)

    # Example: 渡邉理佐・原田葵のスキル発動率が1.8倍
    for m in re.finditer(r"([^\n。]+?)のスキル発動率が\s*([0-9]+(?:\.[0-9]+)?)倍", full):
        val = float(m.group(2))
        names = _parse_member_names_blob(m.group(1))
        for nm in names:
            member_rates[nm] = max(member_rates.get(nm, 1.0), val)

    return color_rates, member_rates


def _parse_vs_mode(name: str) -> str:
    norm = name.replace("、", "・").replace(",", "・")
    if "最も高い効果のみ" in norm:
        return "max_each"
    if "全て合算" in norm:
        return "sum_all"
    if "Vo・Daのみ合算" in norm:
        return "sum_vo_da"
    if "Vo・Peのみ合算" in norm:
        return "sum_vo_pe"
    if "Da・Peのみ合算" in norm:
        return "sum_da_pe"
    if "Voのみ合算" in norm:
        return "sum_vo"
    if "Daのみ合算" in norm:
        return "sum_da"
    if "Peのみ合算" in norm:
        return "sum_pe"
    return "sum_all"


def _mode_zero_axes(mode: str) -> set[str]:
    if mode == "sum_vo":
        return {"da", "pe"}
    if mode == "sum_da":
        return {"vo", "pe"}
    if mode == "sum_pe":
        return {"vo", "da"}
    if mode == "sum_vo_da":
        return {"pe"}
    if mode == "sum_vo_pe":
        return {"da"}
    if mode == "sum_da_pe":
        return {"vo"}
    return set()


def _parse_vs_fixed_bonus(description: str) -> tuple[dict[str, float], set[str]]:
    clean = _strip_tags(description)
    fixed = {a: 0.0 for a in AXES}
    target_types: set[str] = set()
    pattern = re.compile(
        r"(RED|BLUE|GREEN|YELLOW|PURPLE)(?:・(RED|BLUE|GREEN|YELLOW|PURPLE))?"
        r"タイプの([^。]+?)が([0-9]+(?:\.[0-9]+)?)%アップ"
    )
    for m in pattern.finditer(clean):
        colors = {COLOR_WORDS[m.group(1)]}
        if m.group(2):
            colors.add(COLOR_WORDS[m.group(2)])
        axes = _parse_axes_text(m.group(3))
        pct = float(m.group(4))
        target_types.update(colors)
        for a in axes:
            fixed[a] = max(fixed[a], pct)
    return fixed, target_types


def _build_vs_rule(summary_name: str, summary_desc: str, center_color: str) -> VSCenterRule | None:
    if "フロント内のセンタースキルを発動" not in (summary_name or ""):
        return None
    mode = _parse_vs_mode(summary_name)
    m_scale = re.search(r"([0-9]+(?:\.[0-9]+)?)倍で発動", _strip_tags(summary_desc))
    scale = float(m_scale.group(1)) if m_scale else 0.5
    source_types = _extract_colors(summary_name)
    if not source_types:
        source_types = {center_color}
    fixed_bonus, fixed_target_types = _parse_vs_fixed_bonus(summary_desc)
    if not fixed_target_types:
        fixed_target_types = set(source_types)
    return VSCenterRule(
        mode=mode,
        scale=scale,
        source_types=source_types,
        agg_target_types=set(source_types),
        fixed_bonus=fixed_bonus,
        fixed_target_types=fixed_target_types,
    )


def _build_skill_tail_map(skill_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    by_code = {s["code"]: s for s in skill_rows}
    tail: dict[str, dict[str, Any]] = {}
    for code in by_code:
        seen = set()
        cur = by_code[code]
        while cur.get("next_skill_master_code"):
            nxt = cur.get("next_skill_master_code")
            if not nxt or nxt in seen or nxt not in by_code:
                break
            seen.add(nxt)
            cur = by_code[nxt]
        tail[code] = cur
    return tail


def _build_cards(
    masters_dir: Path,
    workbook_path: Path,
    active_name_norms: set[str] | None,
    exclude_name_norms: set[str] | None,
) -> tuple[list[Card], dict[str, Any]]:
    card_rows = _read_json(masters_dir / "card_masters.json")["card_masters"]
    char_rows = _read_json(masters_dir / "character_masters.json")["character_masters"]
    skill_rows = _read_json(masters_dir / "skill_masters.json")["skill_masters"]
    leader_rows = _read_json(masters_dir / "leader_skill_summary_masters.json")["leader_skill_summary_masters"]
    add_rows = _read_json(masters_dir / "additional_card_node_training_masters.json")[
        "additional_card_node_training_masters"
    ]

    expected_map = _excel_skill_expected_map(workbook_path)
    card_after_skill_map = _excel_card_after_skill_map(workbook_path)
    card_scene_upgrade_map = _excel_card_scene_upgrade_total_map(workbook_path)
    chars = {c["code"]: c for c in char_rows}
    leaders = {l["code"]: l for l in leader_rows}
    add_by_card = {a["card_master_code"]: a for a in add_rows}
    skill_tail = _build_skill_tail_map(skill_rows)

    raw_ssr = 0
    active_filtered_out = 0
    excluded_by_manual = 0
    workbook_ssr_plus_fallback_used = 0

    cards: list[Card] = []
    for row in card_rows:
        if row.get("rarity") != 4:
            continue
        raw_ssr += 1

        color = COLOR_BY_FEEL.get(int(row.get("feel_type") or 0))
        if not color:
            continue

        char_code = row["character_master_code"]
        char = chars.get(char_code)
        if not char:
            continue

        member_name = str(char.get("name") or "")
        member_name_norm = _normalize_name(member_name)
        if exclude_name_norms and member_name_norm in exclude_name_norms:
            excluded_by_manual += 1
            continue
        if active_name_norms and member_name_norm not in active_name_norms:
            active_filtered_out += 1
            continue

        add = add_by_card.get(row["code"])
        gain = row.get("total_card_node_gain_parameters") or {}
        ga = int(gain.get("a") or 0)
        gb = int(gain.get("b") or 0)
        gc = int(gain.get("c") or 0)
        title = str(row.get("title") or "")

        # Match the card table/export pipeline:
        #   base = parameter_*_max + total_card_node_gain_parameters.*
        #   plus = additional.parameter_*_max + total_card_node_gain_parameters.* (if SSR+ exists)
        # Use the higher final stat per axis when SSR+ data exists.
        base_vo = int(row.get("parameter_a_max") or 0) + ga
        base_da = int(row.get("parameter_b_max") or 0) + gb
        base_pe = int(row.get("parameter_c_max") or 0) + gc
        if add:
            plus_vo = int(add.get("parameter_a_max") or 0) + ga
            plus_da = int(add.get("parameter_b_max") or 0) + gb
            plus_pe = int(add.get("parameter_c_max") or 0) + gc
            vo = max(base_vo, plus_vo)
            da = max(base_da, plus_da)
            pe = max(base_pe, plus_pe)
        else:
            # Some older cards (notably many disc-era cards) have no
            # additional_card_node_training row in masters, but workbook has
            # an explicit Lv100 scene total (+2250). For those, apply +750 per axis.
            scene_upgrade_total = int(
                card_scene_upgrade_map.get((_title_key(title), member_name_norm, color), 0)
            )
            if scene_upgrade_total >= 2200:
                fallback_bonus = 750
                workbook_ssr_plus_fallback_used += 1
            else:
                # Keep fallback disabled by default unless explicitly configured.
                fallback_bonus = SSR_PLUS_FALLBACK_NODE_BONUS
            vo = base_vo + fallback_bonus
            da = base_da + fallback_bonus
            pe = base_pe + fallback_bonus

        skill_code = row.get("skill_master_code")
        skill_tail_row = skill_tail.get(skill_code) if skill_code else None
        skill_desc = str((skill_tail_row or {}).get("description") or "")
        excel_skill_desc = card_after_skill_map.get((_title_key(title), member_name_norm, color))
        if excel_skill_desc:
            skill_desc = excel_skill_desc
        skill_key = _normalize_skill_key(skill_desc)
        skill_expected = expected_map.get(skill_key)
        if skill_expected is None:
            skill_expected = MANUAL_EXPECTED_OVERRIDES.get(skill_key, 0.0)

        is_score = "スコア" in skill_desc
        is_combo = "コンボボーナス" in skill_desc

        lcode = row.get("leader_skill_summary_master_code")
        lrow = leaders.get(lcode) if lcode else None
        lname = str((lrow or {}).get("name") or "")
        ldesc = str((lrow or {}).get("description") or "")
        leader_effect, _ = _parse_leader_effect(lname)
        leader_skill_rate_color, leader_skill_rate_member = _parse_leader_skill_rate_effects(lname, ldesc)

        title_l = title.lower()
        is_vs_base = ("véaut" in title_l) or ("s.teller" in title_l)
        vs_rule = _build_vs_rule(lname, ldesc, color) if is_vs_base else None

        cards.append(
            Card(
                code=row["code"],
                title=title,
                member_code=char_code,
                member_name=member_name,
                member_name_norm=member_name_norm,
                color=color,
                vo=vo,
                da=da,
                pe=pe,
                skill_desc=skill_desc,
                skill_expected=float(skill_expected or 0.0),
                skill_is_score=is_score,
                skill_is_combo=is_combo,
                leader_name=lname,
                leader_desc=ldesc,
                leader_effect=leader_effect,
                leader_skill_rate_color=leader_skill_rate_color,
                leader_skill_rate_member=leader_skill_rate_member,
                is_vs_base=is_vs_base,
                vs_rule=vs_rule,
            )
        )

    meta = {
        "masters_version": masters_dir.name,
        "all_ssr_before_active_filter": raw_ssr,
        "active_filtered_out": active_filtered_out,
        "excluded_by_manual": excluded_by_manual,
        "all_ssr_count": len(cards),
        "vs_base_count": sum(1 for c in cards if c.is_vs_base and c.vs_rule is not None),
        "active_filter_enabled": bool(active_name_norms),
        "workbook_ssr_plus_fallback_used": workbook_ssr_plus_fallback_used,
    }
    return cards, meta


def _team_included_for_vs(team: list[Card], center: Card, rule: VSCenterRule) -> list[Card]:
    included = [c for c in team if c.color in rule.source_types]
    if center not in included:
        included.insert(0, center)
    if not included:
        return [center]
    return included


def _aggregate_center_bonus(team: list[Card], center: Card, rule: VSCenterRule) -> dict[str, float]:
    included = _team_included_for_vs(team, center, rule)

    vectors: list[dict[str, float]] = []
    for c in included:
        # V/S cards store axis-up in vs_rule.fixed_bonus; regular cards use leader_effect.
        if c.vs_rule is not None:
            v = {a: float(c.vs_rule.fixed_bonus.get(a, 0.0)) for a in AXES}
        else:
            v = {a: float(c.leader_effect.get(a, 0.0)) for a in AXES}
        vectors.append(v)

    agg = {a: 0.0 for a in AXES}

    if rule.mode == "max_each":
        for a in AXES:
            agg[a] = max(v.get(a, 0.0) for v in vectors)
    elif rule.mode == "sum_vo":
        agg["vo"] = sum(v.get("vo", 0.0) for v in vectors)
    elif rule.mode == "sum_da":
        agg["da"] = sum(v.get("da", 0.0) for v in vectors)
    elif rule.mode == "sum_pe":
        agg["pe"] = sum(v.get("pe", 0.0) for v in vectors)
    elif rule.mode == "sum_vo_da":
        agg["vo"] = sum(v.get("vo", 0.0) for v in vectors)
        agg["da"] = sum(v.get("da", 0.0) for v in vectors)
    elif rule.mode == "sum_vo_pe":
        agg["vo"] = sum(v.get("vo", 0.0) for v in vectors)
        agg["pe"] = sum(v.get("pe", 0.0) for v in vectors)
    elif rule.mode == "sum_da_pe":
        agg["da"] = sum(v.get("da", 0.0) for v in vectors)
        agg["pe"] = sum(v.get("pe", 0.0) for v in vectors)
    else:
        for a in AXES:
            agg[a] = sum(v.get(a, 0.0) for v in vectors)

    return {a: agg[a] * rule.scale for a in AXES}


def _compute_effective_stats(team: list[Card], center: Card) -> tuple[int, int, int]:
    rule = center.vs_rule
    if not rule:
        return (
            int(sum(c.vo for c in team)),
            int(sum(c.da for c in team)),
            int(sum(c.pe for c in team)),
        )

    agg = _aggregate_center_bonus(team, center, rule)
    zero_axes = _mode_zero_axes(rule.mode)
    # Game front-detail applies ceil per source-effect contribution, then sums.
    # Aggregated single-pass ceil can drift by a few points on large stats.
    component_pcts: dict[str, list[float]] = {a: [] for a in AXES}
    if rule.mode != "max_each":
        included = _team_included_for_vs(team, center, rule)
        vectors: list[dict[str, float]] = []
        for c in included:
            if c.vs_rule is not None:
                v = {a: float(c.vs_rule.fixed_bonus.get(a, 0.0)) for a in AXES}
            else:
                v = {a: float(c.leader_effect.get(a, 0.0)) for a in AXES}
            vectors.append(v)

        def _append_scaled(axis: str, raw_val: float) -> None:
            if raw_val > 0.0:
                component_pcts[axis].append(raw_val * rule.scale)

        for v in vectors:
            if rule.mode == "sum_vo":
                _append_scaled("vo", float(v.get("vo", 0.0)))
            elif rule.mode == "sum_da":
                _append_scaled("da", float(v.get("da", 0.0)))
            elif rule.mode == "sum_pe":
                _append_scaled("pe", float(v.get("pe", 0.0)))
            elif rule.mode == "sum_vo_da":
                _append_scaled("vo", float(v.get("vo", 0.0)))
                _append_scaled("da", float(v.get("da", 0.0)))
            elif rule.mode == "sum_vo_pe":
                _append_scaled("vo", float(v.get("vo", 0.0)))
                _append_scaled("pe", float(v.get("pe", 0.0)))
            elif rule.mode == "sum_da_pe":
                _append_scaled("da", float(v.get("da", 0.0)))
                _append_scaled("pe", float(v.get("pe", 0.0)))
            else:
                for axis in AXES:
                    _append_scaled(axis, float(v.get(axis, 0.0)))

    total = {"vo": 0, "da": 0, "pe": 0}
    for card in team:
        base = {"vo": float(card.vo), "da": float(card.da), "pe": float(card.pe)}
        if card.color in rule.agg_target_types:
            for a in zero_axes:
                base[a] = 0.0
        for axis in AXES:
            # Non-target colors keep raw scene values.
            if card.color not in rule.agg_target_types:
                total[axis] += int(base[axis])
                continue

            # For max_each, there is only one selected percentage per axis.
            if rule.mode == "max_each":
                total[axis] += int(math.ceil(base[axis] * (1.0 + agg[axis] / 100.0)))
                continue

            # Sum modes: apply ceil for each source contribution then add base.
            subtotal = int(base[axis])
            for pct in component_pcts[axis]:
                subtotal += int(math.ceil(base[axis] * (pct / 100.0)))
            total[axis] += int(subtotal)

    return int(total["vo"]), int(total["da"]), int(total["pe"])


def _collect_skill_rate_multipliers(team: list[Card], center: Card) -> tuple[dict[str, float], dict[str, float]]:
    if center.vs_rule:
        included = _team_included_for_vs(team, center, center.vs_rule)
        rate_scale = center.vs_rule.scale
    else:
        # Non-V/S centers do not chain-absorb teammate skill-rate buffs.
        included = [center]
        rate_scale = 1.0

    color_mult = {c: 1.0 for c in ["R", "B", "G", "Y", "P"]}
    member_mult: dict[str, float] = {}

    def _scaled_rate(m: float) -> float:
        # For V/S absorbed center skills, skill-rate uses additive delta scaling.
        # Example: 2.0x with 0.7 chain => 1 + (2.0-1)*0.7 = 1.7x
        if rate_scale >= 0.999:
            return float(m)
        return 1.0 + max(0.0, float(m) - 1.0) * rate_scale

    for c in included:
        for color, mult in c.leader_skill_rate_color.items():
            color_mult[color] = max(color_mult.get(color, 1.0), _scaled_rate(mult))
        for member, mult in c.leader_skill_rate_member.items():
            member_mult[member] = max(member_mult.get(member, 1.0), _scaled_rate(mult))

    return color_mult, member_mult


def _skill_proc_probability(desc: str) -> float:
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)%\s*の確率", desc or "")
    if not m:
        return 0.20
    p = float(m.group(1)) / 100.0
    return min(0.99, max(0.01, p))


def _team_skill_ev(team: list[Card], center: Card) -> tuple[float, float, float, float, float, float]:
    color_mult, member_mult = _collect_skill_rate_multipliers(team, center)

    score_ev = 0.0
    combo_ev = 0.0
    eff_score_ev = 0.0
    eff_combo_ev = 0.0
    var_score = 0.0
    var_combo = 0.0

    for c in team:
        m_color = color_mult.get(c.color, 1.0)
        m_member = member_mult.get(c.member_name_norm, 1.0)
        m = max(1.0, m_color, m_member)
        p = _skill_proc_probability(c.skill_desc)

        # Keep objective scoring consistent with zawa preset axis behavior:
        # score-like skills contribute to combo-axis, combo-like skills contribute to score-axis.
        if c.skill_is_score:
            combo_ev += c.skill_expected
            e = c.skill_expected * m
            eff_combo_ev += e
            # Bernoulli approximation on EV contribution (for 1σ/2σ interval view).
            var_combo += (e * e) * (1.0 - p) / max(p, 1e-6)
        if c.skill_is_combo:
            score_ev += c.skill_expected
            e = c.skill_expected * m
            eff_score_ev += e
            var_score += (e * e) * (1.0 - p) / max(p, 1e-6)

    score_sigma = math.sqrt(max(var_score, 0.0))
    combo_sigma = math.sqrt(max(var_combo, 0.0))
    return score_ev, combo_ev, eff_score_ev, eff_combo_ev, score_sigma, combo_sigma


def _estimate_score_raw(team_power: int, score_ev_pct: float, combo_ev_pct: float, notes: int) -> int:
    score_bonus = score_ev_pct / 100.0
    combo_bonus = combo_ev_pct / 100.0
    base = team_power / float(max(notes, 1))
    total = 0
    for i in range(1, notes + 1):
        p = i / notes
        val = math.ceil(base * (1 + p * (1 + score_bonus)) * (1 + combo_bonus))
        total += val
    return int(total)


def _axis_weights(rule: VSCenterRule | None) -> dict[str, float]:
    if not rule:
        return {"vo": 1.0, "da": 1.0, "pe": 1.0}
    if rule.mode == "sum_vo":
        return {"vo": 1.4, "da": 0.2, "pe": 0.2}
    if rule.mode == "sum_da":
        return {"vo": 0.2, "da": 1.4, "pe": 0.2}
    if rule.mode == "sum_pe":
        return {"vo": 0.2, "da": 0.2, "pe": 1.4}
    if rule.mode == "sum_vo_da":
        return {"vo": 1.0, "da": 1.0, "pe": 0.2}
    if rule.mode == "sum_vo_pe":
        return {"vo": 1.0, "da": 0.2, "pe": 1.0}
    if rule.mode == "sum_da_pe":
        return {"vo": 0.2, "da": 1.0, "pe": 1.0}
    return {"vo": 1.0, "da": 1.0, "pe": 1.0}


def _objective_for_team(team: list[Card], center: Card) -> TeamResult:
    eff_vo, eff_da, eff_pe = _compute_effective_stats(team, center)
    team_power = eff_vo + eff_da + eff_pe
    score_ev, combo_ev, eff_score_ev, eff_combo_ev, score_sigma, combo_sigma = _team_skill_ev(team, center)

    objective = team_power * (1.0 + eff_score_ev / 300.0) * (1.0 + eff_combo_ev / 100.0)
    bench = _estimate_score_raw(team_power, eff_score_ev, eff_combo_ev, notes=900)

    return TeamResult(
        center=center,
        supports=[c for c in team if c.code != center.code],
        objective=objective,
        bench_score=bench,
        eff_vo=eff_vo,
        eff_da=eff_da,
        eff_pe=eff_pe,
        team_power=team_power,
        score_ev_pct=score_ev,
        combo_ev_pct=combo_ev,
        effective_score_ev_pct=eff_score_ev,
        effective_combo_ev_pct=eff_combo_ev,
        score_sigma_pct=score_sigma,
        combo_sigma_pct=combo_sigma,
    )


def _objective_value(team: list[Card], center: Card) -> float:
    eff_vo, eff_da, eff_pe = _compute_effective_stats(team, center)
    team_power = eff_vo + eff_da + eff_pe
    _, _, eff_score_ev, eff_combo_ev, _, _ = _team_skill_ev(team, center)
    return team_power * (1.0 + eff_score_ev / 300.0) * (1.0 + eff_combo_ev / 100.0)


def _center_focus_axes(rule: VSCenterRule | None) -> set[str]:
    if rule is None:
        return {"vo", "da", "pe"}
    if rule.mode == "sum_vo":
        return {"vo"}
    if rule.mode == "sum_da":
        return {"da"}
    if rule.mode == "sum_pe":
        return {"pe"}
    if rule.mode == "sum_vo_da":
        return {"vo", "da"}
    if rule.mode == "sum_vo_pe":
        return {"vo", "pe"}
    if rule.mode == "sum_da_pe":
        return {"da", "pe"}
    return {"vo", "da", "pe"}


def _leader_axes(card: Card) -> set[str]:
    return {a for a in AXES if float(card.leader_effect.get(a, 0.0)) > 0.0}


def _axis_t1_rank_key(center: Card, card: Card, *, song_color: str = "ALL") -> tuple[int, int, int, float]:
    """
    Strict-priority helper used by app optimizer:
    - T1 first (skill_expected >= 3.0)
    - then same source color(s) + axis-match single-bias
    - then dual-bias
    - then tri-bias
    V-card second color is naturally included via center.vs_rule.source_types.
    """
    source_types = center.vs_rule.source_types if center.vs_rule else set()
    focus_axes = _center_focus_axes(center.vs_rule)
    axes = _leader_axes(card)
    axis_cnt = len(axes) if axes else 3
    axis_match = len(axes & focus_axes)
    song_color_norm = str(song_color or "ALL").strip().upper()
    if song_color_norm not in {"R", "B", "G", "Y", "P"}:
        song_color_norm = "ALL"
    if song_color_norm != "ALL":
        if card.color == song_color_norm:
            color_group = 0  # primary: song color
        elif source_types and card.color in source_types:
            color_group = 1  # secondary: V/S absorbed color
        else:
            color_group = 2
    else:
        color_group = 0 if (source_types and card.color in source_types) else 1
    color_match = 1 if color_group <= 1 else 0
    is_t1 = 1 if float(card.skill_expected) >= 3.0 else 0

    if is_t1:
        if axis_cnt == 1 and axis_match >= 1 and color_group == 0:
            bucket = 0
        elif axis_cnt == 2 and axis_match >= 1 and color_group == 0:
            bucket = 1
        elif axis_cnt == 1 and axis_match >= 1 and color_group == 1:
            bucket = 2
        elif axis_cnt == 2 and axis_match >= 1 and color_group == 1:
            bucket = 3
        elif axis_cnt == 3 and color_group == 0:
            bucket = 4
        elif axis_cnt == 3 and color_group == 1:
            bucket = 5
        elif axis_match >= 1:
            bucket = 6
        else:
            bucket = 7
    else:
        bucket = 8

    # Smaller is better for first three terms. Last term is a tie-break score.
    return (
        bucket,
        color_group,
        0 if axis_match > 0 else 1,
        float(card.skill_expected),
    )


def _support_shortlist(
    center: Card,
    all_cards: list[Card],
    shortlist_size: int,
    candidate_strategy: str = "default",
    song_color: str = "ALL",
) -> list[Card]:
    weights = _axis_weights(center.vs_rule)
    source_types = center.vs_rule.source_types if center.vs_rule else set()
    veaut_member_norms = {c.member_name_norm for c in all_cards if _is_veaut_card(c)}
    center_is_veaut = _is_veaut_card(center)
    # Allow multiple cards from the same member (important for V-card same-name builds).
    pool = [c for c in all_cards if c.code != center.code]

    def score(c: Card) -> float:
        stat = c.vo * weights["vo"] + c.da * weights["da"] + c.pe * weights["pe"]
        skill = c.skill_expected * (1.0 + (1.0 if c.skill_is_combo else 0.2))
        leader = (
            c.leader_effect.get("vo", 0.0) * weights["vo"]
            + c.leader_effect.get("da", 0.0) * weights["da"]
            + c.leader_effect.get("pe", 0.0) * weights["pe"]
        )

        skill_rate_bonus = 0.0
        if c.leader_skill_rate_color:
            skill_rate_bonus += 45.0 * sum(max(v - 1.0, 0.0) for v in c.leader_skill_rate_color.values())
        if c.leader_skill_rate_member:
            skill_rate_bonus += 60.0 * sum(max(v - 1.0, 0.0) for v in c.leader_skill_rate_member.values())

        same_member_bonus = 0.0
        center_member_mult = center.leader_skill_rate_member.get(center.member_name_norm, 1.0)
        if c.member_name_norm == center.member_name_norm and center_member_mult > 1.0:
            same_member_bonus += 240.0 * (center_member_mult - 1.0)
        if center_is_veaut and c.member_name_norm in veaut_member_norms:
            same_member_bonus += 90.0
            if _is_veaut_card(c):
                same_member_bonus += 130.0

        bonus = 1.10 if (source_types and c.color in source_types) else 1.0
        tkey = _title_key(c.title)
        vs_title_bonus = 1.08 if ("veaut" in tkey or "s.teller" in tkey) else 1.0

        return (stat + 120.0 * skill + 80.0 * leader + skill_rate_bonus + same_member_bonus) * bonus * vs_title_bonus

    strategy = (candidate_strategy or "default").strip().lower()
    if strategy == "axis_t1":
        def _key(c: Card) -> tuple[int, int, int, float, float, str]:
            b0, b1, b2, _ = _axis_t1_rank_key(center, c, song_color=song_color)
            return (b0, b1, b2, -score(c), -float(c.skill_expected), c.code)

        ranked = sorted(
            pool,
            key=_key,
        )
    else:
        ranked = sorted(pool, key=score, reverse=True)
    return ranked[:shortlist_size]


def _build_search_pool(
    center: Card,
    all_cards: list[Card],
    shortlist_size: int,
    search_pool_size: int,
    candidate_strategy: str = "default",
    song_color: str = "ALL",
) -> tuple[list[Card], list[Card]]:
    shortlist = _support_shortlist(
        center,
        all_cards,
        shortlist_size,
        candidate_strategy=candidate_strategy,
        song_color=song_color,
    )
    if len(shortlist) < 4:
        return shortlist, shortlist[:]

    ranked = shortlist
    all_pool = [c for c in all_cards if c.code != center.code]

    same_member = [c for c in all_pool if c.member_name_norm == center.member_name_norm]
    same_member = sorted(
        same_member,
        key=lambda c: _objective_value([center, c], center),
        reverse=True,
    )

    vs_related = [c for c in all_pool if c.vs_rule is not None or c.leader_skill_rate_member]
    vs_related = sorted(
        vs_related,
        key=lambda c: _objective_value([center, c], center),
        reverse=True,
    )

    pool: list[Card] = []
    seen: set[str] = set()

    def _push(cards: list[Card], limit: int) -> None:
        for c in cards:
            if c.code in seen:
                continue
            seen.add(c.code)
            pool.append(c)
            if len(pool) >= limit:
                break

    seed_limit = max(8, min(search_pool_size // 3, 16))
    _push(same_member, seed_limit)
    _push(vs_related, max(seed_limit * 2, 20))

    if _is_veaut_card(center):
        veaut_by_member: dict[str, list[Card]] = {}
        for c in all_pool:
            if _is_veaut_card(c):
                veaut_by_member.setdefault(c.member_name_norm, []).append(c)
        veaut_pair_seed: list[Card] = []
        for member, v_cards in veaut_by_member.items():
            v_sorted = sorted(v_cards, key=lambda c: _objective_value([center, c], center), reverse=True)
            veaut_pair_seed.extend(v_sorted[:1])
            non_v_same = [c for c in all_pool if c.member_name_norm == member and not _is_veaut_card(c)]
            non_v_same = sorted(non_v_same, key=lambda c: _objective_value([center, c], center), reverse=True)
            veaut_pair_seed.extend(non_v_same[:2])
        veaut_pair_seed = sorted(veaut_pair_seed, key=lambda c: _objective_value([center, c], center), reverse=True)
        _push(veaut_pair_seed, max(seed_limit * 3, 28))

    _push(ranked, search_pool_size)

    if len(pool) < 4:
        pool = ranked[: max(4, search_pool_size)]
    return ranked, pool


def _build_team_candidates(
    center: Card,
    all_cards: list[Card],
    shortlist_size: int,
    search_pool_size: int,
    topk: int,
    candidate_strategy: str = "default",
    song_color: str = "ALL",
) -> list[TeamResult]:
    ranked, pool = _build_search_pool(
        center,
        all_cards,
        shortlist_size,
        search_pool_size,
        candidate_strategy=candidate_strategy,
        song_color=song_color,
    )
    if len(pool) < 4:
        team = [center] + pool[:4]
        return [_objective_for_team(team, center)]

    scored: list[tuple[float, list[Card]]] = []
    for combo in itertools.combinations(pool, 4):
        team = [center, *combo]
        obj = _objective_value(team, center)
        scored.append((obj, team))
    scored.sort(key=lambda x: x[0], reverse=True)

    out: list[TeamResult] = []
    seen: set[tuple[str, ...]] = set()
    want = max(1, int(topk))
    for _obj, team in scored:
        key = tuple(sorted(c.code for c in team))
        if key in seen:
            continue
        seen.add(key)
        out.append(_objective_for_team(team, center))
        if len(out) >= want:
            break

    if not out:
        team = [center] + ranked[:4]
        out = [_objective_for_team(team, center)]
    return out


def _build_team(center: Card, all_cards: list[Card], shortlist_size: int, search_pool_size: int) -> TeamResult:
    return _build_team_candidates(
        center=center,
        all_cards=all_cards,
        shortlist_size=shortlist_size,
        search_pool_size=search_pool_size,
        topk=1,
    )[0]


def _parse_duration_seconds(text: str) -> int:
    m = re.search(r"(\d+):(\d+)", text or "")
    if not m:
        return 0
    return int(m.group(1)) * 60 + int(m.group(2))


def _parse_int_text(text: str) -> int:
    m = re.search(r"(\d+)", text or "")
    return int(m.group(1)) if m else 0


def _parse_multiplier(text: str) -> float:
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", text or "")
    return float(m.group(1)) if m else 1.0


def _fetch_songlist() -> dict[str, Any]:
    js_code = _http_get_text(UNIAIR_SONGLIST_URL)
    try:
        import quickjs  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise SystemExit(
            "quickjs is required to refresh songlist. "
            "Install in .venv: .venv/bin/pip install quickjs\n"
            f"details: {exc}"
        )

    ctx = quickjs.Context()
    ctx.eval(js_code)
    raw = ctx.eval("JSON.stringify(song)")
    rows = json.loads(raw)

    songs: list[dict[str, Any]] = []
    for row in rows:
        color_raw = str(row.get("color") or "").strip().upper()
        color = SONG_COLOR_WORDS.get(color_raw, "ALL")
        seconds = _parse_duration_seconds(str(row.get("time") or ""))
        notes = _parse_int_text(str(row.get("notes") or ""))
        level = _parse_int_text(str(row.get("Lv") or ""))
        psylli = _parse_multiplier(str(row.get("psylli") or "1.0"))
        if notes <= 0 or seconds <= 0:
            continue
        songs.append(
            {
                "no": int(row.get("no") or 0),
                "color": color,
                "name": str(row.get("songs") or ""),
                "live": str(row.get("live") or ""),
                "level": level,
                "seconds": seconds,
                "notes": notes,
                "psylli": psylli,
            }
        )

    return {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source": UNIAIR_SONGLIST_URL,
        "count": len(songs),
        "songs": songs,
    }


def _load_songlist(path: Path, refresh: bool) -> list[Song]:
    if refresh or not path.exists():
        data = _fetch_songlist()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    if not path.exists():
        return []

    data = _read_json(path)
    rows = data.get("songs", [])
    out: list[Song] = []
    for r in rows:
        try:
            out.append(
                Song(
                    no=int(r.get("no") or 0),
                    color=str(r.get("color") or "ALL"),
                    name=str(r.get("name") or ""),
                    live=str(r.get("live") or ""),
                    level=int(r.get("level") or 0),
                    seconds=int(r.get("seconds") or 0),
                    notes=int(r.get("notes") or 0),
                    psylli=float(r.get("psylli") or 1.0),
                )
            )
        except Exception:
            continue
    return [s for s in out if s.notes > 0 and s.seconds > 0]


def _load_song_allowlist(path: Path) -> dict[str, set[str]]:
    if not path.exists():
        raise SystemExit(f"Song allowlist not found: {path}")

    rows: list[tuple[str, str]] = []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise SystemExit(f"Song allowlist csv has no header: {path}")
            color_key = "song_color" if "song_color" in reader.fieldnames else "color"
            name_key = "song_name" if "song_name" in reader.fieldnames else "name"
            for r in reader:
                color = str((r.get(color_key) or "")).strip().upper()
                name = str((r.get(name_key) or "")).strip()
                if color and name:
                    rows.append((color, name))
    else:
        data = _read_json(path)
        src = data.get("songs", data)
        if isinstance(src, list):
            for x in src:
                if not isinstance(x, dict):
                    continue
                color = str((x.get("song_color") or x.get("color") or "")).strip().upper()
                name = str((x.get("song_name") or x.get("name") or "")).strip()
                if color and name:
                    rows.append((color, name))

    out: dict[str, set[str]] = {}
    for color, name in rows:
        mapped = SONG_COLOR_WORDS.get(color, color)
        out.setdefault(mapped, set()).add(_normalize_song_name(name))
    return out


def _filter_songs_by_allowlist(songs: list[Song], allow: dict[str, set[str]]) -> list[Song]:
    if not allow:
        return songs

    out: list[Song] = []
    for s in songs:
        sname = _normalize_song_name(s.name)
        # Keep color matching strict. "ALL" list is only for songs whose own color is ALL.
        allow_names = set(allow.get(s.color, set()))
        if s.color == "ALL":
            allow_names.update(allow.get("ALL", set()))
        if sname in allow_names:
            out.append(s)
    return out


def _estimate_song_score(
    team: TeamResult,
    song: Song,
    group_power: int,
    member_point: int,
    song_scale: float,
) -> tuple[int, float, float, int, int, int, int]:
    def _scaled_score(score_ev: float, combo_ev: float) -> int:
        raw_inner = _estimate_score_raw(team.team_power, score_ev, combo_ev, song.notes)
        gp_ratio = group_power / float(max(team.team_power, 1))
        member_factor = 1.0 + (member_point / 150000.0)
        density = song.notes / float(max(song.seconds, 1))
        density_factor = 1.0 + min(density / 12.0, 1.0) * 0.35
        return int(round(raw_inner * gp_ratio * song.psylli * member_factor * density_factor * song_scale))

    score_ev = team.effective_score_ev_pct
    combo_ev = team.effective_combo_ev_pct
    scaled = _scaled_score(score_ev, combo_ev)
    score_million = scaled / 1_000_000.0

    # Delta-method style local sensitivity + Gaussian bands for 1σ/2σ.
    eps_s = max(0.01, team.score_sigma_pct / 50.0)
    eps_c = max(0.01, team.combo_sigma_pct / 50.0)
    dsd = (_scaled_score(score_ev + eps_s, combo_ev) - _scaled_score(max(0.0, score_ev - eps_s), combo_ev)) / (
        2.0 * eps_s
    )
    dcd = (_scaled_score(score_ev, combo_ev + eps_c) - _scaled_score(score_ev, max(0.0, combo_ev - eps_c))) / (
        2.0 * eps_c
    )
    sigma = math.sqrt((dsd * team.score_sigma_pct) ** 2 + (dcd * team.combo_sigma_pct) ** 2)
    sigma_i = int(round(sigma))
    s1_low = max(0, int(round(scaled - sigma)))
    s1_high = int(round(scaled + sigma))
    s2_low = max(0, int(round(scaled - 2.0 * sigma)))
    s2_high = int(round(scaled + 2.0 * sigma))
    return scaled, score_million, sigma, s1_low, s1_high, s2_low, s2_high


def _axis_label(axis: str) -> str:
    if axis == "vo":
        return "Vo"
    if axis == "da":
        return "Da"
    return "Pe"


def _team_effect_lines(team: TeamResult) -> list[str]:
    center = team.center
    rule = center.vs_rule
    if not rule:
        return []

    lines: list[str] = []
    agg = _aggregate_center_bonus(team.cards, center, rule)
    for color in sorted(rule.agg_target_types):
        for axis in AXES:
            pct = agg.get(axis, 0.0)
            if pct > 0.05:
                lines.append(f"{color}タイプの{_axis_label(axis)}が{pct:.1f}%アップ")

    zero_axes = _mode_zero_axes(rule.mode)
    if zero_axes:
        kept = [a for a in AXES if a not in zero_axes]
        kept_text = "・".join(_axis_label(a) for a in kept)
        zero_text = "・".join(_axis_label(a) for a in sorted(zero_axes))
        lines.append(f"{kept_text}のみ合算（{zero_text}を0として計算）")

    _, member_mult = _collect_skill_rate_multipliers(team.cards, center)
    seen_members: set[str] = set()
    for c in team.cards:
        key = c.member_name_norm
        if key in seen_members:
            continue
        seen_members.add(key)
        m = member_mult.get(key, 1.0)
        if m > 1.0001:
            lines.append(f"{c.member_name}のスキル発動率が{(m - 1.0) * 100.0:.1f}%アップ")
    return lines


def _front_total(team: TeamResult, member_point: int) -> tuple[int, int, int]:
    member_total = member_point * len(team.cards)
    scene_total = team.team_power
    return scene_total, member_total, scene_total + member_total


def _rank_teams_by_song_color(
    teams: list[TeamResult],
    songs: list[Song],
    group_power: int,
    member_point: int,
    song_scale: float,
) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {}
    for song_color in ("G", "Y", "R", "P", "B"):
        color_songs = [s for s in songs if s.color == song_color]
        if not color_songs:
            continue
        rows: list[dict[str, Any]] = []
        for t in teams:
            scores: list[int] = []
            sigmas: list[float] = []
            for s in color_songs:
                score_raw, _, sigma, _, _, _, _ = _estimate_song_score(t, s, group_power, member_point, song_scale)
                scores.append(score_raw)
                sigmas.append(float(sigma))
            scene_total, member_total, front_total = _front_total(t, member_point)
            rows.append(
                {
                    "song_color": song_color,
                    "avg_score": int(round(sum(scores) / float(len(scores)))),
                    "median_score": int(round(statistics.median(scores))),
                    "avg_sigma": int(round(sum(sigmas) / float(len(sigmas)))),
                    "center_member": t.center.member_name,
                    "center_title": t.center.title,
                    "center_color": t.center.color,
                    "team_power_scene": scene_total,
                    "member_point_total": member_total,
                    "front_total": front_total,
                    "eff_score_ev_pct": round(t.effective_score_ev_pct, 3),
                    "eff_combo_ev_pct": round(t.effective_combo_ev_pct, 3),
                    "effect_summary": " / ".join(_team_effect_lines(t)),
                    "team_cards": " | ".join(f"{c.member_name}[{c.title}]" for c in t.cards),
                }
            )
        rows.sort(key=lambda r: (r["avg_score"], r["median_score"], -r["avg_sigma"]), reverse=True)
        for i, r in enumerate(rows, start=1):
            r["rank"] = i
        out[song_color] = rows
    return out


def _write_outputs(
    out_dir: Path,
    results: list[TeamResult],
    songs: list[Song],
    meta: dict[str, Any],
    group_power: int,
    member_point: int,
    song_scale: float,
    song_allowlist_size: int,
    enable_approx_color_top5: bool,
) -> tuple[Path, Path, Path, Path, Path, Path | None, Path | None, Path | None, Path | None]:
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "vs_base_all_ssr_teams.csv"
    md_path = out_dir / "vs_base_all_ssr_teams.md"
    song_csv = out_dir / "vs_base_song_scores_all.csv"
    song_top15_csv = out_dir / "vs_base_song_scores_top15.csv"
    song_top15_md = out_dir / "vs_base_song_scores_top15.md"
    color_top5_csv = out_dir / "vs_base_color_top5_teams.csv"
    color_top5_md = out_dir / "vs_base_color_top5_teams.md"
    color_top5_center_only_csv = out_dir / "vs_base_color_top5_teams_center_color_only.csv"
    color_top5_center_only_md = out_dir / "vs_base_color_top5_teams_center_color_only.md"

    sorted_results = sorted(results, key=lambda x: x.bench_score, reverse=True)

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "rank",
                "center_card",
                "center_member",
                "center_color",
                "center_mode",
                "center_source_types",
                "bench_score",
                "objective",
                "team_power",
                "eff_vo",
                "eff_da",
                "eff_pe",
                "score_ev_pct",
                "combo_ev_pct",
                "effective_score_ev_pct",
                "effective_combo_ev_pct",
                "score_sigma_pct",
                "combo_sigma_pct",
                "team_cards",
            ]
        )
        for i, r in enumerate(sorted_results, start=1):
            mode = r.center.vs_rule.mode if r.center.vs_rule else ""
            source = "".join(sorted(r.center.vs_rule.source_types)) if r.center.vs_rule else ""
            team_cards = " | ".join(f"{c.member_name}[{c.title}]" for c in r.cards)
            w.writerow(
                [
                    i,
                    r.center.title,
                    r.center.member_name,
                    r.center.color,
                    mode,
                    source,
                    r.bench_score,
                    f"{r.objective:.2f}",
                    r.team_power,
                    r.eff_vo,
                    r.eff_da,
                    r.eff_pe,
                    f"{r.score_ev_pct:.2f}",
                    f"{r.combo_ev_pct:.2f}",
                    f"{r.effective_score_ev_pct:.2f}",
                    f"{r.effective_combo_ev_pct:.2f}",
                    f"{r.score_sigma_pct:.2f}",
                    f"{r.combo_sigma_pct:.2f}",
                    team_cards,
                ]
            )

    by_color: dict[str, list[TeamResult]] = {}
    for r in results:
        by_color.setdefault(r.center.color, []).append(r)

    lines: list[str] = []
    lines.append("# V/S Base + All SSR Accessory Teams")
    lines.append("")
    lines.append(f"- masters_version: `{meta['masters_version']}`")
    lines.append(f"- all_ssr_before_active_filter: `{meta['all_ssr_before_active_filter']}`")
    lines.append(f"- active_filtered_out: `{meta['active_filtered_out']}`")
    lines.append(f"- excluded_by_manual: `{meta.get('excluded_by_manual', 0)}`")
    lines.append(f"- all_ssr_count: `{meta['all_ssr_count']}`")
    lines.append(f"- vs_base_count: `{meta['vs_base_count']}`")
    lines.append(f"- active_filter_enabled: `{meta['active_filter_enabled']}`")
    lines.append("- score model objective: `team_power * (1 + eff_score_ev/300) * (1 + eff_combo_ev/100)`")
    lines.append("- eff_score_ev / eff_combo_ev include skill-rate multipliers from center-skill chain")
    lines.append("- bench score: notes=900 deterministic estimate")
    lines.append("")

    for color in ("R", "B", "G", "Y", "P"):
        color_rows = sorted(by_color.get(color, []), key=lambda x: x.bench_score, reverse=True)[:5]
        if not color_rows:
            continue
        lines.append(f"## {color} Top 5")
        lines.append("")
        for i, r in enumerate(color_rows, start=1):
            mode = r.center.vs_rule.mode if r.center.vs_rule else "-"
            source = ",".join(sorted(r.center.vs_rule.source_types)) if r.center.vs_rule else "-"
            lines.append(
                f"{i}. `{r.center.member_name}[{r.center.title}]` "
                f"bench=`{r.bench_score}` power=`{r.team_power}` "
                f"(Vo {r.eff_vo} / Da {r.eff_da} / Pe {r.eff_pe}) "
                f"EV(base)=`score {r.score_ev_pct:.2f}% + combo {r.combo_ev_pct:.2f}%` "
                f"EV(eff)=`score {r.effective_score_ev_pct:.2f}% + combo {r.effective_combo_ev_pct:.2f}%` "
                f"SIGMA=`score {r.score_sigma_pct:.2f}% + combo {r.combo_sigma_pct:.2f}%` "
                f"mode=`{mode}` source=`{source}`"
            )
            lines.append("   " + " | ".join(f"{c.member_name}[{c.title}]" for c in r.cards))
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")

    # pick best team per color + global best for ALL songs
    best_team_by_color: dict[str, TeamResult] = {}
    for color in ("R", "B", "G", "Y", "P"):
        rows = sorted(by_color.get(color, []), key=lambda x: x.bench_score, reverse=True)
        if rows:
            best_team_by_color[color] = rows[0]
    if sorted_results:
        best_team_by_color["ALL"] = sorted_results[0]

    song_rows: list[dict[str, Any]] = []
    for song in songs:
        team = best_team_by_color.get(song.color) or best_team_by_color.get("ALL")
        if not team:
            continue
        score_raw, score_m, sigma, s1_low, s1_high, s2_low, s2_high = _estimate_song_score(
            team, song, group_power, member_point, song_scale
        )
        song_rows.append(
            {
                "song_no": song.no,
                "song_color": song.color,
                "song_name": song.name,
                "live": song.live,
                "level": song.level,
                "seconds": song.seconds,
                "notes": song.notes,
                "psylli": song.psylli,
                "team_center": f"{team.center.member_name}[{team.center.title}]",
                "team_cards": " | ".join(f"{c.member_name}[{c.title}]" for c in team.cards),
                "team_power": team.team_power,
                "eff_score_ev_pct": round(team.effective_score_ev_pct, 3),
                "eff_combo_ev_pct": round(team.effective_combo_ev_pct, 3),
                "score_sigma_pct": round(team.score_sigma_pct, 3),
                "combo_sigma_pct": round(team.combo_sigma_pct, 3),
                "estimated_score": score_raw,
                "estimated_score_million": round(score_m, 3),
                "score_sigma": int(round(sigma)),
                "score_1sigma_low": s1_low,
                "score_1sigma_high": s1_high,
                "score_2sigma_low": s2_low,
                "score_2sigma_high": s2_high,
            }
        )

    with song_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(song_rows[0].keys()) if song_rows else ["song_no"])
        w.writeheader()
        for row in sorted(song_rows, key=lambda x: (x["song_color"], -x["estimated_score"])):
            w.writerow(row)

    top15_rows: list[dict[str, Any]] = []
    for color in ("ALL", "R", "B", "G", "Y", "P"):
        rows = [r for r in song_rows if r["song_color"] == color]
        rows = sorted(rows, key=lambda x: x["estimated_score"], reverse=True)[:15]
        for i, r in enumerate(rows, start=1):
            row = dict(r)
            row["rank_in_color"] = i
            top15_rows.append(row)

    with song_top15_csv.open("w", newline="", encoding="utf-8") as f:
        cols = [
            "song_color",
            "rank_in_color",
            "song_no",
            "song_name",
            "live",
            "level",
            "seconds",
            "notes",
            "psylli",
            "estimated_score",
            "estimated_score_million",
            "score_sigma",
            "score_1sigma_low",
            "score_1sigma_high",
            "score_2sigma_low",
            "score_2sigma_high",
            "team_center",
            "team_power",
            "eff_score_ev_pct",
            "eff_combo_ev_pct",
            "score_sigma_pct",
            "combo_sigma_pct",
        ]
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for row in top15_rows:
            w.writerow({k: row.get(k, "") for k in cols})

    md_lines: list[str] = []
    md_lines.append("# Song Scores (Top 15 Per Color)")
    md_lines.append("")
    md_lines.append(f"- group_power: `{group_power}`")
    md_lines.append(f"- member_point: `{member_point}`")
    md_lines.append(f"- song_scale: `{song_scale}`")
    md_lines.append(f"- song_allowlist_size: `{song_allowlist_size}`")
    md_lines.append("")
    for color in ("ALL", "R", "B", "G", "Y", "P"):
        rows = [r for r in top15_rows if r["song_color"] == color]
        if not rows:
            continue
        md_lines.append(f"## {color}")
        md_lines.append("")
        for r in rows:
            md_lines.append(
                f"{r['rank_in_color']}. `{r['song_name']}` "
                f"score=`{r['estimated_score']}` ({r['estimated_score_million']}M) "
                f"σ=`{r['score_sigma']}` "
                f"1σ=[`{r['score_1sigma_low']}`, `{r['score_1sigma_high']}`] "
                f"2σ=[`{r['score_2sigma_low']}`, `{r['score_2sigma_high']}`] "
                f"notes=`{r['notes']}` time=`{r['seconds']}s` "
                f"team=`{r['team_center']}`"
            )
        md_lines.append("")

    song_top15_md.write_text("\n".join(md_lines), encoding="utf-8")

    color_top5_csv_out: Path | None = None
    color_top5_md_out: Path | None = None
    color_top5_center_only_csv_out: Path | None = None
    color_top5_center_only_md_out: Path | None = None

    if songs and enable_approx_color_top5:
        all_ranked = _rank_teams_by_song_color(results, songs, group_power, member_point, song_scale)
        center_only_ranked = _rank_teams_by_song_color(
            [r for r in results if r.center.color in {"G", "Y", "R", "P", "B"}],
            songs,
            group_power,
            member_point,
            song_scale,
        )

        def _write_color_top5(
            ranked: dict[str, list[dict[str, Any]]],
            out_csv: Path,
            out_md: Path,
            title: str,
            center_color_only: bool,
        ) -> tuple[Path, Path]:
            csv_rows: list[dict[str, Any]] = []
            md_lines_local: list[str] = [title, ""]
            for color in ("G", "Y", "R", "P", "B"):
                rows = ranked.get(color, [])
                if center_color_only:
                    rows = [r for r in rows if r["center_color"] == color]
                rows = rows[:5]
                if not rows:
                    continue
                md_lines_local.append(f"## {color}")
                md_lines_local.append("")
                for idx, r in enumerate(rows, start=1):
                    row = dict(r)
                    row["rank"] = idx
                    csv_rows.append(row)
                    md_lines_local.append(
                        f"{idx}. `{r['center_member']}[{r['center_title']}]` "
                        f"avg=`{r['avg_score']}` median=`{r['median_score']}` sigma=`{r['avg_sigma']}` "
                        f"フロント総合力=`{r['front_total']}` "
                        f"(scene {r['team_power_scene']} + member {r['member_point_total']}) "
                        f"EV=`{r['eff_score_ev_pct']}% + {r['eff_combo_ev_pct']}%`"
                    )
                    md_lines_local.append(f"   {r['team_cards']}")
                    if r["effect_summary"]:
                        md_lines_local.append(f"   発動する効果: {r['effect_summary']}")
                md_lines_local.append("")

            cols = [
                "song_color",
                "rank",
                "avg_score",
                "median_score",
                "avg_sigma",
                "center_member",
                "center_title",
                "center_color",
                "team_power_scene",
                "member_point_total",
                "front_total",
                "eff_score_ev_pct",
                "eff_combo_ev_pct",
                "effect_summary",
                "team_cards",
            ]
            with out_csv.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=cols)
                w.writeheader()
                for row in csv_rows:
                    w.writerow({k: row.get(k, "") for k in cols})
            out_md.write_text("\n".join(md_lines_local), encoding="utf-8")
            return out_csv, out_md

        color_top5_csv_out, color_top5_md_out = _write_color_top5(
            all_ranked,
            color_top5_csv,
            color_top5_md,
            "# Top5 Teams By Song Color (Avg score over fixed 15 songs)",
            center_color_only=False,
        )
        color_top5_center_only_csv_out, color_top5_center_only_md_out = _write_color_top5(
            center_only_ranked,
            color_top5_center_only_csv,
            color_top5_center_only_md,
            "# Top5 Teams By Song Color (Center color == song color)",
            center_color_only=True,
        )

    return (
        csv_path,
        md_path,
        song_csv,
        song_top15_csv,
        song_top15_md,
        color_top5_csv_out,
        color_top5_md_out,
        color_top5_center_only_csv_out,
        color_top5_center_only_md_out,
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Optimize V/S base teams with all SSR cards as accessories."
    )
    parser.add_argument(
        "--masters-dir",
        type=Path,
        default=None,
        help="masters/<version> directory. Default: latest under ./masters",
    )
    parser.add_argument(
        "--masters-root",
        type=Path,
        default=Path("masters"),
        help="masters root (used when --masters-dir is omitted)",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("UOA大表 新人必看.xlsx"),
        help="Workbook path for skill expected mapping",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        required=True,
        help="Output directory (suggested: update/<RUN_TS>_masters_<VERSION>/tables)",
    )
    parser.add_argument(
        "--shortlist-size",
        type=int,
        default=240,
        help="Support shortlist size per center",
    )
    parser.add_argument(
        "--search-pool-size",
        type=int,
        default=36,
        help="Exhaustive search pool size per center (choose-4 on this pool)",
    )
    parser.add_argument(
        "--active-members-json",
        type=Path,
        default=Path("catalogs/active_members_wiki.json"),
        help="Cache file for active members from Japanese Wikipedia",
    )
    parser.add_argument(
        "--refresh-active-members",
        action="store_true",
        help="Refresh active members cache from Wikipedia before run",
    )
    parser.add_argument(
        "--no-active-filter",
        action="store_true",
        help="Disable active member filter",
    )
    parser.add_argument(
        "--songlist-json",
        type=Path,
        default=Path("catalogs/uniair_songlist.json"),
        help="Cache file for UNI'AIR songlist (from uniair_sim songlist.js)",
    )
    parser.add_argument(
        "--refresh-songlist",
        action="store_true",
        help="Refresh songlist cache from uniair_sim before run (requires quickjs)",
    )
    parser.add_argument(
        "--no-song-scores",
        action="store_true",
        help="Skip song-by-song score export",
    )
    parser.add_argument(
        "--group-power",
        type=int,
        default=1_800_000,
        help="Assumed group power for song score scaling",
    )
    parser.add_argument(
        "--member-point",
        type=int,
        default=15_000,
        help="Assumed per-member point for song score scaling",
    )
    parser.add_argument(
        "--song-scale",
        type=float,
        default=4.5,
        help="Global scaling factor for song score output",
    )
    parser.add_argument(
        "--song-allowlist",
        type=Path,
        default=None,
        help="Optional song allowlist (.csv/.json). If set, only these songs are scored.",
    )
    parser.add_argument(
        "--exclude-members",
        type=str,
        default="",
        help="Comma-separated member names to exclude explicitly (after normalization).",
    )
    parser.add_argument(
        "--enable-approx-color-top5",
        action="store_true",
        help="Enable legacy approximate color-top5 ranking export (disabled by default).",
    )
    args = parser.parse_args()

    masters_dir = args.masters_dir or _find_latest_masters(args.masters_root)

    active_name_norms: set[str] | None = None
    if not args.no_active_filter:
        active_name_norms = _load_active_members(args.active_members_json, args.refresh_active_members)

    exclude_name_norms: set[str] | None = None
    if args.exclude_members.strip():
        names = re.split(r"[,\n;|/]+", args.exclude_members)
        exclude_name_norms = {_normalize_name(x) for x in names if _normalize_name(x)}

    songs: list[Song] = []
    if not args.no_song_scores:
        songs = _load_songlist(args.songlist_json, args.refresh_songlist)
        if args.song_allowlist:
            allow = _load_song_allowlist(args.song_allowlist)
            songs = _filter_songs_by_allowlist(songs, allow)

    cards, meta = _build_cards(masters_dir, args.workbook, active_name_norms, exclude_name_norms)
    base_centers = [c for c in cards if c.is_vs_base and c.vs_rule is not None]
    if not base_centers:
        raise SystemExit("No V/S base cards found from the given masters.")

    results: list[TeamResult] = []
    for center in base_centers:
        res = _build_team(center, cards, args.shortlist_size, args.search_pool_size)
        results.append(res)

    (
        csv_path,
        md_path,
        song_csv,
        song_top15_csv,
        song_top15_md,
        color_top5_csv,
        color_top5_md,
        color_top5_center_only_csv,
        color_top5_center_only_md,
    ) = _write_outputs(
        args.out_dir,
        results,
        songs,
        meta,
        group_power=args.group_power,
        member_point=args.member_point,
        song_scale=args.song_scale,
        song_allowlist_size=len(songs),
        enable_approx_color_top5=args.enable_approx_color_top5,
    )

    print(f"masters_version={meta['masters_version']}")
    print(f"all_ssr_before_active_filter={meta['all_ssr_before_active_filter']}")
    print(f"active_filtered_out={meta['active_filtered_out']}")
    print(f"excluded_by_manual={meta.get('excluded_by_manual', 0)}")
    print(f"all_ssr_count={meta['all_ssr_count']}")
    print(f"vs_base_count={meta['vs_base_count']}")
    print(f"active_filter_enabled={meta['active_filter_enabled']}")
    print(f"results={len(results)}")
    print(f"csv={csv_path}")
    print(f"md={md_path}")
    if songs:
        print(f"songs={len(songs)}")
        print(f"song_csv={song_csv}")
        print(f"song_top15_csv={song_top15_csv}")
        print(f"song_top15_md={song_top15_md}")
        if color_top5_csv and color_top5_md:
            print(f"color_top5_csv={color_top5_csv}")
            print(f"color_top5_md={color_top5_md}")
        elif not args.enable_approx_color_top5:
            print("color_top5=skipped (approx model disabled; use strict zawa rerank flow)")
        if color_top5_center_only_csv and color_top5_center_only_md:
            print(f"color_top5_center_only_csv={color_top5_center_only_csv}")
            print(f"color_top5_center_only_md={color_top5_center_only_md}")


if __name__ == "__main__":
    main()
